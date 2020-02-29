# download the transaction zipped file. Extract and paste in this project folder
# run: pipenv install openpyxl
#change your virtualenv to PyExcel
# import openpyxl
# call the load_workbook() method on openpyxl & pass the "transactions.xlsx" as value in the method  & save in a var wb
# print(wb.sheetnames) run the code
# run: wb["Sheet1"] save it in sheet var
# run: sheet["a1"] save in cell var
# run: for row in range(1, sheet.max_row + 1):
# nest another for loop: for column in range(1, sheet.max_column + 1)
# under same code block run: sheet.cell(row, column) save in cell var
# print(cell.value) 
# also after sheet["a1"] you can run the following:
# sheet["a"] this will return all the cells in column a. store it in column var & print
# you can also run: sheet["a1:c3"]
# we also have: sheet.append([1,2,3]) to add to the tables
# finally run: wb.save("transactions2.xlsx") this would be used to save the file name

import openpyxl
wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet1"]
cell = sheet["a1"]
# for row in range(1, sheet.max_row + 1):
#     for column in range(1, sheet.max_column + 1):
#         cell = sheet.cell(row, column)
#         print(cell.value)
column = sheet["a"]
column = sheet["a1:c3"]
print(column)

sheet.append([1,2,3])
wb.save("transactions2.xlsx")